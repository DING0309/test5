from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "Agenda"

# 定义数据
data = [
    ("09:0009:30", "報到", "無"),
    ("09:3009:40", "開場致詞", "王教授"),
    ("09:4010:05", "邁向6G 的AI-RAN及O-RAN 趨勢介紹", "劉教授"),
    ("10:0510:30", "下世代B5G/6G專網應用與未來趨勢", "陳教授"),
    ("10:3010:50", "Break", "無"),
    ("10:5011:20", "從O-RAN到AI-RAN 智慧通訊的節能應用", "教學團隊"),
    ("11:2012:00", "O-RAN環境和各模組化功能介紹", "教學團隊"),
    ("12:0013:30", "Lunch", "無"),
    ("13:3014:00", "O-RAN 的市場應用案例", "教學團隊"),
    ("14:0014:30", "O-RAN OSC環境建置教學", "教學團隊"),
    ("14:3014:50", "Break", "無"),
    ("14:5015:50", "O-RAN OSC第三方應用程式 xApps建置教學", "教學團隊"),
    ("15:5016:30", "現場討論時間", "教學團隊"),
]

# 添加表头
headers = ["時間", "內容", "講者"]
ws.append(headers)

# 添加数据并处理合并
for i, (time_str, content, speaker) in enumerate(data, start=2):  # 从第2行开始
    # 格式化时间
    start_time = time_str[:5]  # 前5字符，如09:00
    end_time = time_str[5:]    # 后5字符，如09:30
    formatted_time = f"{start_time}~{end_time}"
    
    ws.cell(row=i, column=1, value=formatted_time)
    
    if speaker == "無":
        # 合并内容和讲者列
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
        ws.cell(row=i, column=2, value=content)
    else:
        ws.cell(row=i, column=2, value=content)
        ws.cell(row=i, column=3, value=speaker)

# 定义边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 定义居中+自动换行对齐样式
center_alignment = Alignment(
    horizontal='center',
    vertical='center',
    wrap_text=True
)

# 应用样式到所有单元格
for row in ws.iter_rows(min_row=1, max_row=len(data) + 1, min_col=1, max_col=3):
    for cell in row:
        cell.alignment = center_alignment
        cell.border = thin_border

# 设置列宽
ws.column_dimensions['A'].width = 15  # 時間列
ws.column_dimensions['B'].width = 40  # 內容列
ws.column_dimensions['C'].width = 15  # 講者列

# 设置行高以适应换行文本
for row in ws.iter_rows(min_row=1, max_row=len(data) + 1):
    ws.row_dimensions[row[0].row].height = 30

# 保存文件
wb.save('agenda.xlsx')
print("agenda.xlsx 已成功生成！")